from flask import Flask
app = Flask(__name__)
from openpyxl import load_workbook

@app.route('/')
def menuMaker():
	region='ANTIOQUIA'
	area='BELLAS ARTES'
	percentile=14
	try:
		wb = load_workbook('Options_Outcomes.xlsx')
		ws = wb.get_sheet_by_name('Sheet1')
		results=len(ws['a'])
	except: results="NO FUNCIONA EL EXCEL"
	
	return results



if __name__ == '__main__':
  app.run()
